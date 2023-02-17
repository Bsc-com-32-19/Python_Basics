from typing import Any

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1, 1)

    # number of rows
    for row in range(1, sheet.max_row + 1):
        print(row)
    # print cell from second roe we don't need the title
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
    corrected_price: float | Any = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)

    # set the value for this cell
    corrected_price_cell.value = corrected_price
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       max_col=4,
                       min_col=4)

    # creating an instance of BarChart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)



